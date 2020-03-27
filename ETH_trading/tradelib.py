import pandas as pd
import numpy as np
import shelve
import datetime

from openpyxl import load_workbook

from app import client, user_data

""" Constants """

max_open_risk = 0.05  # 5%
commission = 0.001  # 0.1%

""" Dictionaries for dropdowns etc """

pairs = [
    {'label': 'ETH/USDT', 'value': 'ETHUSDT'},
    {'label': 'BTC/USDT', 'value': 'BTCUSDT'},
    {'label': 'XRP/USDT', 'value': 'XRPUSDT'},
         ]

types = [
    {'label': 'pullback to value', 'value': 'pullback to value'},
    {'label': 'ATR extreme', 'value': 'ATR extreme'},
    {'label': 'price rejection', 'value': 'price rejection'}
]

directions = [
    {'label': 'LONG', 'value': 'LONG'},
    {'label': 'SHORT', 'value': 'SHORT'}
]

spans = [
    {'label': 'Daily', 'value': 'D'},
    {'label': 'Weekly', 'value': 'W'},
    {'label': 'Monthly', 'value': 'M'}
]

open_trade_cols = ['pair', 'size', 'entry', 'stop', 'direction']
open_trade_dict = [{'name': c, 'id': c} for c in open_trade_cols]

closed_trade_cols = ['pair', 'size', 'entry', 'stop', 'exit', 'P/L (%)',
                     'risk (%)', 'RRR', 'cap. share (%)', 'timespan (min)', 'direction', 'type', 'confidence', 'note']
closed_trade_dict = [{'name': c, 'id': c} for c in closed_trade_cols]


""" Functions """


def max_qty(entry, stop, max_risk, leverage=1):
    """ Calculate the allowed position size

    Note: Works for both SHORT and LONG trades. Be careful to use the right input order!

    If user chooses to trade with leverage, position size is limited to 5x capital
    """

    cap = user_data['capital'][-1]
    max_cap_risk = -cap*max_risk
    qty_cap = cap/entry*leverage                 # max quantity limited by capital

    # max quantity by risk:
    if entry > stop:    # LONG
        qty_risk = max_cap_risk/(stop*(1-commission)**2-entry)
    else:               # SHORT
        qty_risk = max_cap_risk/(entry*(1-commission) - stop*(1+commission))

    return min(qty_risk, qty_cap)


def trade_risk(cap, trades):
    """ Calculate the risk (%) of (an array of) open trades

    INPUT: cap is the current capital, trades is a df with !open! trades
    """

    entry = trades['entry'].to_numpy()
    size = trades['size'].to_numpy()
    stop = trades['stop'].to_numpy()

    high = np.maximum(entry, stop)
    low = np.minimum(entry, stop)

    # NOTE: This formula is an approximation! Now we can use one equation for short and long trades.
    loss = size*abs(high/(1-commission) - low*(1-commission))
    risk = loss/cap*100

    return risk


def open_profit(trades):

    # TODO: This for loop makes calls to binance API, so execution speed depends on how many calls per second binance
    #  allows us to make. Better to save a current price somewhere that is updated every x seconds.

    current_price = pd.Series([])
    for s in trades['pair']:
        ticker = client.get_symbol_ticker(symbol=s)
        cp = float(ticker['price'])
        current_price = current_price.append(pd.Series(cp), ignore_index=True)

    size = trades['size']
    entry = trades['entry']

    open_value = size.multiply(current_price*(1-commission) - entry)

    return sum(open_value)


def profit_abs(trade):
    t = trade.iloc[0]
    entry = t['entry']
    close = t['exit']
    qty = t['size']
    direction = t['direction']

    if direction == 'SHORT':
        return qty*(entry*(1-commission) - close*(1+commission))
    elif direction == 'LONG':
        return qty*(close*(1-commission)**2 - entry)
    else:
        raise NameError('I do not recognize the trade direction "' + direction + '"')


def profit_rel(trade):
    cap = user_data['capital'][-1]
    p_abs = profit_abs(trade)
    p_rel = p_abs/cap*100
    return p_rel


def risk_reward_ratio(trade):
    t = trade.iloc[0]
    entry = t['entry']
    stop = t['stop']
    close = t['exit']
    if close <= stop and t['direction'] == 'LONG':
        return 0
    elif close >= stop and t['direction'] == 'SHORT':
        return 0
    else:
        return abs((close-entry)/(entry-stop))


def capital_target(n_days, month_profit_target):
    """ Calculate the evolution of the capital over time with a constant profit each day. This is a target that a
    trader can try to aim for.

    INPUTS: month_profit - target profit (%) for a month
            n_days - number of days to predict forwards
    """
    start_date = datetime.datetime.date(user_data['capital'].index[0])
    d = pd.date_range(start=start_date, periods=n_days, freq='D')

    # Calculate the capital over time with a constant profit each day:
    start_cap = user_data['capital'][0]
    daily_profit = (1+month_profit_target/100)**(12/365.25)-1
    v = np.ones(n_days)*(1+daily_profit)
    cap_array = np.cumprod(v)*start_cap/(1+daily_profit)

    prediction = pd.Series(cap_array, d)

    return prediction


def update_shelf(var_name, new_data):
    """" Update a variable in the shelf as a moving average

    INPUT: new_point is a pd.Series with datetimeIndex
    """
    # with shelve.open('user_data') as d:
    # Retrieve the old win rate and amount of recorded points:
    data = user_data[var_name]
    old_avg = data[-1]
    n = len(data)
    # Calculate the new win rate:
    new_avg = (n*old_avg + new_data)/(n+1)
    # Add the result to the user_data:
    user_data[var_name] = data.append(new_avg)

    return new_avg


def update_win_rate(trade):
    t = trade.iloc[0]
    is_winner = t['P/L (%)'] > 0

    df = pd.Series(int(is_winner), trade.index)
    new_win_rate = update_shelf('win_rate', df)

    return new_win_rate


def update_avg_profit(trade):
    t = trade.iloc[0]
    p_rel = t['P/L (%)']

    df = pd.Series(p_rel, trade.index)
    new_avg_profit = update_shelf('avg_profit', df)

    return new_avg_profit


def update_avg_rrr(trade):
    t = trade.iloc[0]
    rrr = risk_reward_ratio(trade)
    if rrr == 0:  # trade was a loss, no rrr update
        new_avg_rrr = user_data['avg_rrr'][-1]
    else:
        df = pd.Series(rrr, trade.index)
        new_avg_rrr = update_shelf('avg_rrr', df)

    return new_avg_rrr


def update_avg_timespan(trade):
    t = trade.iloc[0]
    delta = t['timespan (min)']

    df = pd.Series(delta, trade.index)
    new_avg_timespan = update_shelf('avg_timespan', df)

    return new_avg_timespan


def update_total_profit(trade):
    p = profit_abs(trade)  # p_abs means the absolute profit in $, can be negative!
    if p < 0:
        user_data['total_loss'] += p
        user_data['n_losses'] += 1
    else:
        user_data['total_gain'] += p
        user_data['n_wins'] += 1


def update_expectancy(trade):
    # NOTE: This function MUST be called AFTER update_total_profit.
    # If you don't do this you update the expectancy with an old trade
    wins = user_data['n_wins']
    losses = user_data['n_losses']

    if losses == 0:
        avg_loss = 0
    else:
        avg_loss = user_data['total_loss'] / losses

    if wins == 0:
        avg_win = 0
    else:
        avg_win = user_data['total_gain'] / wins

    win_rate = user_data['win_rate'][-1]
    expectancy = avg_win*win_rate + (1-win_rate)*avg_loss

    df = pd.Series(expectancy, trade.index)
    data = user_data['expectancy']
    user_data['expectancy'] = data.append(df)

    return expectancy


def update_capital(trade):
    cap_array = user_data['capital']
    p = profit_abs(trade)
    new_cap = cap_array[-1] + p

    df = pd.Series(new_cap, trade.index)
    user_data['capital'] = cap_array.append(df)

    return new_cap


def update_user_data(trade):
    update_win_rate(trade)
    update_total_profit(trade)
    update_avg_profit(trade)
    update_avg_rrr(trade)
    update_avg_timespan(trade)
    update_expectancy(trade)
    update_capital(trade)
    user_data['n_trades'] += 1
    user_data.sync()


def fill_trade(open_trade, close, note='-'):
    """" Fill in the remaining values from an open trade into a closed trade dataFrame"""
    t = open_trade.iloc[0]

    index = pd.DatetimeIndex([datetime.datetime.now()])
    closed_trade = pd.DataFrame(index=index, columns=closed_trade_cols)
    closed_trade['note'] = note
    closed_trade['exit'] = close

    # Copy values that already exist (except date because we will use the timespan):
    for c in open_trade.drop(columns=['date']).columns:
        closed_trade[c] = t[c]
    # Fill in the remaining values:
    entry = t['entry']
    qty = t['size']
    cap = user_data['capital'][-1]

    closed_trade['P/L (%)'] = profit_rel(closed_trade)
    closed_trade['risk (%)'] = trade_risk(cap, closed_trade)
    closed_trade['RRR'] = risk_reward_ratio(closed_trade)
    closed_trade['cap. share (%)'] = entry*qty/cap*100
    dt = datetime.datetime.now() - t['date']
    minutes = round(dt.seconds / 60)
    closed_trade['timespan (min)'] = minutes

    closed_trade = closed_trade.round(2)
    return closed_trade


def read_trades(record_file, status, dict_output=False):
    trades = pd.read_excel(record_file, sheet_name=status)
    if dict_output:
        trades = trades.drop(columns=['date'])
        # For using this function in a dash table we need a dict as output:
        trades = trades.to_dict(orient='records')
    return trades


def write_trade_to_records(record_file, status, trade):
    with pd.ExcelWriter(path=record_file, engine='openpyxl', datetime_format='DD-MM-YYYY hh:mm', mode='a') as \
            writer:
        # Open the file:
        writer.book = load_workbook(record_file)
        # Copy existing sheets:
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        # Add new trade on top of the existing data:
        sheet = status
        writer.book[sheet].insert_rows(2)
        trade.to_excel(writer, sheet_name=sheet, startrow=1, header=None, index_label='date')
        writer.close()


def remove_trade_from_records(record_file, idx):
    records = load_workbook(record_file)
    open_sheet = records['open']
    open_sheet.delete_rows(idx+2)
    records.save(record_file)
    # ! THIS FUNCTION OVERWRITES THE RECORDS, DANGEROUS !


if __name__ == '__main__':
    tr = pd.read_excel('diary.xlsx', sheet_name='closed')

    # tr['date'] = tr['date'].apply(lambda x: datetime.datetime.date(x))

    table_data = tr.loc[:, ['date', 'P/L (%)']]
    table_data['P/L (%)'] = table_data['P/L (%)'] / 100 + 1

    group = table_data.groupby(pd.Grouper(key='date', freq='D'))
    merged = group.prod()
    merged['count'] = group.count()

    merged['period'] = merged.index
    merged['period'] = merged['period'].apply(lambda x: datetime.datetime.strftime(x, '%Y %b'))

    # x = tr.iloc[[0]]
    #
    # update_user_data(x)
    # trades = pd.read_excel('diary.xlsx', sheet_name='open')
    #
    # close = 204
    # note = 'tradeLib test'
    #
    # closed_trade = fill_trade(trades, close, note)
    #
    # write_trade_to_records('diary.xlsx', 'closed', closed_trade)
    # remove_trade_from_records('diary.xlsx', 0)

    # import reset_shelf
    # user_data = shelve.open('user_data')
    # cap = user_data['capital']
    #
    # wl_rate = user_data['win_rate']
    # x = trades.tail(1)
    #
    # risk = trade_risk(cap, x)
    #
    # y = fill_trade(x, 201)
    #
    # pred = capital_target(200)
    #
    # p = user_data['avg_profit']
    # wr = user_data['win_rate']
    # r = user_data['avg_rrr']


